from __future__ import annotations

import os
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook, load_workbook

# 正式分頁名稱
TRAVEL_SUBMIT_SHEET = "申請表單"
TRAVEL_DRAFT_SHEET = "草稿列表"

# Canonical column order for Travel
TRAVEL_COLUMNS: List[str] = [
    "id",
    "status",
    "filler_name",
    "form_date",
    "traveler_name",
    "employee_no",
    "plan_code",
    "purpose_desc",
    "travel_route",
    "start_time",
    "end_time",
    "travel_days",
    "is_gov_car",
    "gov_car_no",
    "is_taxi",
    "is_private_car",
    "private_car_km",
    "private_car_no",
    "is_dispatch_car",
    "is_hsr",
    "is_airplane",
    "is_other_transport",
    "other_transport_desc",
    "estimated_cost",
    "expense_rows",
    "total_amount",
    "handler_name",
    "project_manager_name",
    "dept_manager_name",
    "accountant_name",
    "attachments",
    "created_at",
    "updated_at",
    "submitted_at",
]

TRAVEL_COLUMNS_ZH_MAP = {
    "id": "表單編號",
    "status": "狀態",
    "filler_name": "填表人",
    "form_date": "填表日期",
    "traveler_name": "出差人",
    "employee_no": "員工編號",
    "plan_code": "計畫編號",
    "purpose_desc": "出差事由",
    "travel_route": "出差行程",
    "start_time": "出差起始時間",
    "end_time": "出差結束時間",
    "travel_days": "出差天數",
    "is_gov_car": "公務車",
    "gov_car_no": "公務車號",
    "is_taxi": "計程車",
    "is_private_car": "私車",
    "private_car_km": "私車公里數",
    "private_car_no": "私車車號",
    "is_dispatch_car": "派車",
    "is_hsr": "高鐵",
    "is_airplane": "飛機",
    "is_other_transport": "其他交通",
    "other_transport_desc": "其他交通說明",
    "estimated_cost": "預估總花費",
    "expense_rows": "出差明細(JSON)",
    "total_amount": "總金額",
    "handler_name": "經手人",
    "project_manager_name": "計畫主持人",
    "dept_manager_name": "部門主管",
    "accountant_name": "會計",
    "attachments": "附件",
    "created_at": "建立時間",
    "updated_at": "更新時間",
    "submitted_at": "送出時間",
}


def _ensure_parent_dir(xlsx_path: str) -> None:
    parent = os.path.dirname(xlsx_path)
    if parent:
        os.makedirs(parent, exist_ok=True)


def _header_zh() -> List[str]:
    return [TRAVEL_COLUMNS_ZH_MAP.get(c, c) for c in TRAVEL_COLUMNS]


def _create_sheet_with_dual_header(wb, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    ws.append(TRAVEL_COLUMNS)
    ws.append(_header_zh())
    return ws


def ensure_workbook(xlsx_path: str, sheet_name: str) -> None:
    """
    確保活頁簿存在，且指定分頁存在。
    分頁格式：
    - 第1列英文欄名
    - 第2列中文欄名
    - 第3列起資料
    """
    _ensure_parent_dir(xlsx_path)

    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(TRAVEL_COLUMNS)
        ws.append(_header_zh())
        wb.save(xlsx_path)
        return

    cleanup_old_sheets(xlsx_path)

    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        _create_sheet_with_dual_header(wb, sheet_name)
        wb.save(xlsx_path)
    wb.close()


def _read_df(xlsx_path: str, sheet_name: str) -> pd.DataFrame:
    ensure_workbook(xlsx_path, sheet_name)

    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str, header=[0, 1])
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
    except Exception:
        try:
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str)
        except Exception:
            ensure_workbook(xlsx_path, sheet_name)
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str)

    for c in TRAVEL_COLUMNS:
        if c not in df.columns:
            df[c] = ""

    df = df[TRAVEL_COLUMNS].fillna("")

    # 若第二列中文表頭被當成資料，過濾掉
    if not df.empty and "id" in df.columns:
        df = df[df["id"].astype(str).str.strip() != TRAVEL_COLUMNS_ZH_MAP["id"]]

    return df.reset_index(drop=True)


def _write_df_dual_header(xlsx_path: str, df: pd.DataFrame, sheet_name: str) -> None:
    """
    以雙表頭寫入：
    第1列英文 / 第2列中文 / 第3列起資料
    """
    _ensure_parent_dir(xlsx_path)

    df = df.copy()
    for c in TRAVEL_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    df = df[TRAVEL_COLUMNS].fillna("")

    zh_df = pd.DataFrame([_header_zh()], columns=TRAVEL_COLUMNS)
    df_save = pd.concat([zh_df, df], ignore_index=True)

    if os.path.exists(xlsx_path):
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_save.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
    else:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="w") as writer:
            df_save.to_excel(writer, sheet_name=sheet_name, index=False, header=True)


def load_all_travel(
    xlsx_path: str,
    draft_sheet: str = TRAVEL_DRAFT_SHEET,
    submit_sheet: str = TRAVEL_SUBMIT_SHEET,
) -> pd.DataFrame:
    """
    讀取草稿 + 送出兩張分頁並合併。
    """
    ensure_workbook(xlsx_path, draft_sheet)
    ensure_workbook(xlsx_path, submit_sheet)

    dfs = []
    for sn in [draft_sheet, submit_sheet]:
        try:
            df = _read_df(xlsx_path, sn)
            df["_sheet_name"] = sn
            dfs.append(df)
        except Exception:
            pass

    if not dfs:
        return pd.DataFrame(columns=TRAVEL_COLUMNS + ["_sheet_name"])

    df_all = pd.concat(dfs, ignore_index=True)
    if not df_all.empty and "id" in df_all.columns:
        df_all = df_all.sort_values("id", ascending=False, kind="mergesort").reset_index(drop=True)
    return df_all


def upsert_travel_record(xlsx_path: str, record: Dict, sheet_name: str) -> None:
    """
    在指定分頁 upsert 一筆 travel record。
    """
    df = _read_df(xlsx_path, sheet_name)
    rid = str(record.get("id", "")).strip()
    if not rid:
        raise ValueError("record.id is required")

    row = {c: str(record.get(c, "")) for c in TRAVEL_COLUMNS}

    hit = df["id"].astype(str).str.strip() == rid
    if hit.any():
        df.loc[hit, TRAVEL_COLUMNS] = pd.DataFrame([row])[TRAVEL_COLUMNS].values
    else:
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    df = df.sort_values("id", ascending=False, kind="mergesort").reset_index(drop=True)
    _write_df_dual_header(xlsx_path, df, sheet_name)


def delete_travel_record(xlsx_path: str, record_id: str, sheet_name: str) -> None:
    df = _read_df(xlsx_path, sheet_name)
    rid = str(record_id).strip()
    df = df[df["id"].astype(str).str.strip() != rid]
    df = df.sort_values("id", ascending=False, kind="mergesort").reset_index(drop=True)
    _write_df_dual_header(xlsx_path, df, sheet_name)


def cleanup_old_sheets(xlsx_path: str) -> None:
    """
    將舊分頁名稱遷移為新正式名稱，並清掉不再使用的分頁。
    """
    if not os.path.exists(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path)
        changed = False

        rename_map = {
            "DomesticTrip": TRAVEL_SUBMIT_SHEET,
            "DomesticTrip_Draft": TRAVEL_DRAFT_SHEET,
            "出差申請單": TRAVEL_SUBMIT_SHEET,
            "出差草稿": TRAVEL_DRAFT_SHEET,
        }

        for old_name, new_name in rename_map.items():
            if old_name in wb.sheetnames:
                if new_name not in wb.sheetnames:
                    wb[old_name].title = new_name
                    changed = True
                else:
                    old_ws = wb[old_name]
                    new_ws = wb[new_name]
                    # 如果新分頁基本上沒有資料，就用舊分頁取代
                    if new_ws.max_row <= 2:
                        del wb[new_name]
                        old_ws.title = new_name
                        changed = True
                    else:
                        # 新分頁已有資料，刪除舊分頁避免衝突
                        del wb[old_name]
                        changed = True

        obsolete_names = [
            "vouchers",
            "DomesticTrip",
            "DomesticTrip_Draft",
            "出差申請單",
            "出差草稿",
        ]
        for old_name in obsolete_names:
            if old_name in wb.sheetnames:
                del wb[old_name]
                changed = True

        if not wb.sheetnames:
            ws = wb.create_sheet(TRAVEL_SUBMIT_SHEET)
            ws.append(TRAVEL_COLUMNS)
            ws.append(_header_zh())
            changed = True

        if changed:
            wb.save(xlsx_path)
        wb.close()
    except Exception:
        pass